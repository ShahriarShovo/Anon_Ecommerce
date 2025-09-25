from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.db.models import Sum, Count
from django.utils import timezone
from datetime import timedelta, datetime
from django.db.models.functions import TruncDate, TruncDay
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def sales_analytics(request):
    """
    Get sales analytics for daily, weekly, and monthly periods.
    """
    try:
        from orders.models.orders.order import Order
        
        today = timezone.localdate()
        
        # Daily Sales
        daily_sales_data = Order.objects.filter(
            created_at__date=today,
            status='delivered'
        ).aggregate(
            total_amount=Sum('total_amount'),
            order_count=Count('id')
        )
        daily_sales = {
            'total_amount': float(daily_sales_data['total_amount'] or 0),
            'order_count': daily_sales_data['order_count'] or 0,
            'date': today.isoformat()
        }

        # Weekly Sales (last 7 days including today)
        seven_days_ago = today - timedelta(days=6)
        weekly_sales_data = Order.objects.filter(
            created_at__date__range=[seven_days_ago, today],
            status='delivered'
        ).aggregate(
            total_amount=Sum('total_amount'),
            order_count=Count('id')
        )
        weekly_sales = {
            'total_amount': float(weekly_sales_data['total_amount'] or 0),
            'order_count': weekly_sales_data['order_count'] or 0,
            'period': '7 days'
        }

        # Monthly Sales (last 30 days including today)
        thirty_days_ago = today - timedelta(days=29)
        monthly_sales_data = Order.objects.filter(
            created_at__date__range=[thirty_days_ago, today],
            status='delivered'
        ).aggregate(
            total_amount=Sum('total_amount'),
            order_count=Count('id')
        )
        monthly_sales = {
            'total_amount': float(monthly_sales_data['total_amount'] or 0),
            'order_count': monthly_sales_data['order_count'] or 0,
            'period': '30 days'
        }

        # Sales trend for the last 7 days (for chart)
        chart_data = []
        for i in range(7):
            day = today - timedelta(days=6 - i)
            day_sales = Order.objects.filter(
                created_at__date=day,
                status='delivered'
            ).aggregate(
                total_amount=Sum('total_amount'),
                order_count=Count('id')
            )
            chart_data.append({
                'date': day.isoformat(),
                'day_name': day.strftime('%a'), # Abbreviated weekday name
                'total_amount': float(day_sales['total_amount'] or 0),
                'order_count': day_sales['order_count'] or 0
            })
        
        # Sales breakdown by status
        status_breakdown = Order.objects.values('status').annotate(
            total_amount=Sum('total_amount'),
            order_count=Count('id')
        ).order_by('status')
        
        status_breakdown_list = [
            {
                'status': item['status'],
                'total_amount': float(item['total_amount'] or 0),
                'order_count': item['order_count'] or 0
            } for item in status_breakdown
        ]

        response_data = {
            'daily_sales': daily_sales,
            'weekly_sales': weekly_sales,
            'monthly_sales': monthly_sales,
            'chart_data': chart_data,
            'status_breakdown': status_breakdown_list
        }

        return Response({
            'success': True,
            'data': response_data
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({
            'success': False,
            'message': f'Failed to fetch sales analytics: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def sales_trends(request):
    """
    Get sales trends for different time periods.
    """
    try:
        from orders.models.orders.order import Order
        
        today = timezone.localdate()
        
        # Get trends for different periods
        periods = {
            'today': today,
            'yesterday': today - timedelta(days=1),
            'this_week': today - timedelta(days=7),
            'last_week': today - timedelta(days=14),
            'this_month': today - timedelta(days=30),
            'last_month': today - timedelta(days=60)
        }
        
        trends_data = {}
        
        for period_name, start_date in periods.items():
            if period_name in ['today', 'yesterday']:
                # Single day
                end_date = start_date
            elif period_name in ['this_week', 'last_week']:
                # 7 days
                end_date = start_date + timedelta(days=6)
            else:
                # 30 days
                end_date = start_date + timedelta(days=29)
            
            sales_data = Order.objects.filter(
                created_at__date__range=[start_date, end_date],
                status='delivered'
            ).aggregate(
                total_amount=Sum('total_amount'),
                order_count=Count('id')
            )
            
            trends_data[period_name] = {
                'total_amount': float(sales_data['total_amount'] or 0),
                'order_count': sales_data['order_count'] or 0,
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat()
            }
        
        return Response({
            'success': True,
            'data': trends_data
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({
            'success': False,
            'message': f'Failed to fetch sales trends: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def generate_excel_report(request):
    """
    Generate Excel report for sales data
    """
    try:
        from orders.models.orders.order import Order
        from accounts.models import User
        from products.models.products.product import Product
        
        # Get parameters
        report_type = request.GET.get('type', 'sales')  # sales, orders, users, products
        period = request.GET.get('period', '30')  # 7, 30, 365, custom
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        # Calculate date range
        today = timezone.now().date()
        
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            if period == '7':
                start_date = today - timedelta(days=7)
                end_date = today
            elif period == '30':
                start_date = today - timedelta(days=30)
                end_date = today
            elif period == '365':
                start_date = today - timedelta(days=365)
                end_date = today
            else:
                start_date = today - timedelta(days=30)
                end_date = today
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if report_type == 'sales':
            # Sales Report
            ws = wb.active
            ws.title = "Sales Report"
            
            # Headers
            headers = ['Date', 'Orders', 'Total Amount', 'Average Order Value', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Get sales data
            orders = Order.objects.filter(
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            ).order_by('created_at')
            
            # Group by date
            daily_sales = {}
            for order in orders:
                date = order.created_at.date()
                if date not in daily_sales:
                    daily_sales[date] = {
                        'orders': 0,
                        'total_amount': 0,
                        'statuses': {}
                    }
                daily_sales[date]['orders'] += 1
                daily_sales[date]['total_amount'] += float(order.total_amount)
                daily_sales[date]['statuses'][order.status] = daily_sales[date]['statuses'].get(order.status, 0) + 1
            
            # Add data rows
            row = 2
            for date in sorted(daily_sales.keys()):
                data = daily_sales[date]
                avg_order_value = data['total_amount'] / data['orders'] if data['orders'] > 0 else 0
                status_summary = ', '.join([f"{status}: {count}" for status, count in data['statuses'].items()])
                
                ws.cell(row=row, column=1, value=date.strftime('%Y-%m-%d')).border = border
                ws.cell(row=row, column=2, value=data['orders']).border = border
                ws.cell(row=row, column=3, value=f"${data['total_amount']:.2f}").border = border
                ws.cell(row=row, column=4, value=f"${avg_order_value:.2f}").border = border
                ws.cell(row=row, column=5, value=status_summary).border = border
                row += 1
            
            # Add summary row
            total_orders = sum(data['orders'] for data in daily_sales.values())
            total_amount = sum(data['total_amount'] for data in daily_sales.values())
            avg_order_value = total_amount / total_orders if total_orders > 0 else 0
            
            ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=row, column=2, value=total_orders).font = Font(bold=True)
            ws.cell(row=row, column=3, value=f"${total_amount:.2f}").font = Font(bold=True)
            ws.cell(row=row, column=4, value=f"${avg_order_value:.2f}").font = Font(bold=True)
            ws.cell(row=row, column=5, value="").font = Font(bold=True)
            
        elif report_type == 'orders':
            # Orders Report
            ws = wb.active
            ws.title = "Orders Report"
            
            # Headers
            headers = ['Order ID', 'Customer', 'Date', 'Status', 'Total Amount', 'Payment Status', 'Items']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Get orders data
            orders = Order.objects.filter(
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            ).select_related('user', 'payment').prefetch_related('items').order_by('-created_at')
            
            # Add data rows
            row = 2
            for order in orders:
                items = ', '.join([f"{item.product.title} x{item.quantity}" for item in order.items.all()])
                payment_status = order.payment.status if order.payment else 'No Payment'
                
                ws.cell(row=row, column=1, value=order.id).border = border
                ws.cell(row=row, column=2, value=order.user.email).border = border
                ws.cell(row=row, column=3, value=order.created_at.strftime('%Y-%m-%d %H:%M')).border = border
                ws.cell(row=row, column=4, value=order.status).border = border
                ws.cell(row=row, column=5, value=f"${order.total_amount:.2f}").border = border
                ws.cell(row=row, column=6, value=payment_status).border = border
                ws.cell(row=row, column=7, value=items).border = border
                row += 1
                
        elif report_type == 'users':
            # Users Report
            ws = wb.active
            ws.title = "Users Report"
            
            # Headers
            headers = ['User ID', 'Email', 'Full Name', 'Date Joined', 'Last Login', 'Status', 'Orders Count', 'Total Spent']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Get users data
            users = User.objects.filter(
                date_joined__date__gte=start_date,
                date_joined__date__lte=end_date
            ).select_related('profile').prefetch_related('orders')
            
            # Add data rows
            row = 2
            for user in users:
                orders_count = user.orders.count()
                total_spent = sum(float(order.total_amount) for order in user.orders.all())
                
                ws.cell(row=row, column=1, value=user.id).border = border
                ws.cell(row=row, column=2, value=user.email).border = border
                ws.cell(row=row, column=3, value=user.profile.full_name if user.profile else '').border = border
                ws.cell(row=row, column=4, value=user.date_joined.strftime('%Y-%m-%d')).border = border
                ws.cell(row=row, column=5, value=user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never').border = border
                ws.cell(row=row, column=6, value='Active' if user.is_active else 'Inactive').border = border
                ws.cell(row=row, column=7, value=orders_count).border = border
                ws.cell(row=row, column=8, value=f"${total_spent:.2f}").border = border
                row += 1
                
        elif report_type == 'products':
            # Products Report
            ws = wb.active
            ws.title = "Products Report"
            
            # Headers
            headers = ['Product ID', 'Title', 'Category', 'Price', 'Status', 'Created Date', 'Orders Count', 'Total Sold']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Get products data
            products = Product.objects.filter(
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            ).select_related('category').prefetch_related('order_items')
            
            # Add data rows
            row = 2
            for product in products:
                orders_count = product.order_items.count()
                total_sold = sum(item.quantity for item in product.order_items.all())
                
                price_display = f"${product.price:.2f}" if product.price is not None else "N/A"
                created_date = product.created_at.strftime('%Y-%m-%d') if product.created_at else "N/A"
                
                ws.cell(row=row, column=1, value=product.id).border = border
                ws.cell(row=row, column=2, value=product.title).border = border
                ws.cell(row=row, column=3, value=product.category.name if product.category else '').border = border
                ws.cell(row=row, column=4, value=price_display).border = border
                ws.cell(row=row, column=5, value=product.status).border = border
                ws.cell(row=row, column=6, value=created_date).border = border
                ws.cell(row=row, column=7, value=orders_count).border = border
                ws.cell(row=row, column=8, value=total_sold).border = border
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        filename = f"{report_type}_report_{start_date}_to_{end_date}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        import traceback
        # print(f"Excel Report Error: {str(e)}")
        # print(f"Traceback: {traceback.format_exc()}")
        return Response({
            'success': False,
            'message': f'Failed to generate Excel report: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)